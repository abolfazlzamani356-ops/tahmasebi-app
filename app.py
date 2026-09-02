import os
import io
import json
import jdatetime
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify

from models import (
    db, Shop, Category, User, Settings, Customer, BankAccount,
    InventoryItem, StockLog, StockTransfer,
    Invoice, InvoiceItem, Cheque, SalarySlip,
    PettyCashDeposit, Expense, AuditLog
)
from helpers import (
    PERSIAN_MONTHS, DEFAULT_CATEGORIES, RETURN_REASONS,
    get_current_shamsi, log_activity, record_stock_change,
    calculate_seller_exact_stats, get_or_create_customer,
    parse_smart_invoice_text, get_inventory_ai_insights
)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'tahmasebi-mega-erp-v14-permanent-secure-2026')

# رمز نجات مدیریت
MASTER_ADMIN_PASSWORD = os.environ.get('MASTER_ADMIN_PASSWORD', 'king68abolfazl@68')

# مسیر ذخیره دیتابیس پایدار
DATA_DIR = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH', '/data')
if not os.path.exists(DATA_DIR):
    DATA_DIR = os.path.join(app.root_path, 'instance')
    os.makedirs(DATA_DIR, exist_ok=True)

db_path = os.path.join(DATA_DIR, 'tahmasebi_store_persistent.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# ==================== احراز هویت و دسترسی ====================
@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('seller_dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username, is_active=True).first()
        
        is_admin_master = (user and user.role == 'admin' and password == MASTER_ADMIN_PASSWORD)
        
        if user and (user.check_password(password) or is_admin_master):
            session['user_id'] = user.id
            session['full_name'] = user.full_name
            session['role'] = user.role
            session['shop_id'] = user.shop_id or 1
            log_activity("ورود به سامانه" + (" (با رمز نجات)" if is_admin_master else ""), user.full_name, "امنیت")
            return redirect(url_for('index'))
        else:
            flash('نام کاربری یا رمز عبور اشتباه است.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    name = session.get('full_name', 'کاربر')
    log_activity("خروج از سامانه", name, "امنیت")
    session.clear()
    return redirect(url_for('login'))

# ==================== پنل فروشنده ====================
@app.route('/seller')
def seller_dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    now_j = jdatetime.datetime.now()
    selected_month = request.args.get('month', default=now_j.month, type=int)
    user = User.query.get(session['user_id'])
    settings = Settings.query.first()
    
    stats = calculate_seller_exact_stats(user.id, now_j.year, selected_month, user.commission_rate, settings)
    
    invoices = Invoice.query.filter(
        (Invoice.seller_id == user.id) | (Invoice.second_seller_id == user.id),
        Invoice.shamsi_year == now_j.year,
        Invoice.shamsi_month == selected_month
    ).order_by(Invoice.created_at.desc()).all()
    
    colleagues = User.query.filter(User.id != user.id, User.is_active == True).all()
    other_shops = Shop.query.filter(Shop.id != user.shop_id).all()
    inventory_items = InventoryItem.query.filter_by(shop_id=user.shop_id).all()
    all_categories = Category.query.all()
    
    all_sellers = User.query.filter_by(role='seller', is_active=True).all()
    bank_accounts = BankAccount.query.filter_by(is_active=True).all()
    leaderboard = []
    for s in all_sellers:
        s_stats = calculate_seller_exact_stats(s.id, now_j.year, selected_month, s.commission_rate, settings)
        leaderboard.append({'user': s, 'net_sales': s_stats['net_sales'], 'avg_rating': s_stats['avg_rating']})
    leaderboard.sort(key=lambda x: x['net_sales'], reverse=True)

    return render_template(
        'seller_dashboard.html',
        user=user,
        stats=stats,
        invoices=invoices,
        months=PERSIAN_MONTHS,
        selected_month=selected_month,
        current_month_name=PERSIAN_MONTHS.get(selected_month, ''),
        current_year=now_j.year,
        all_categories=all_categories,
        return_reasons=RETURN_REASONS,
        colleagues=colleagues,
        other_shops=other_shops,
        inventory_items=inventory_items,
        bank_accounts=bank_accounts,
        leaderboard=leaderboard
    )

# ==================== صدور و مدیریت فاکتور متصل به انبار (با پرداخت ترکیبی) ====================
@app.route('/invoice/add', methods=['POST'])
def add_invoice():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    now_j = jdatetime.datetime.now()
    doc_status = request.form.get('doc_status', 'final')
    inv_type = 'return' if doc_status == 'return' else 'sale'
    status = 'proforma' if doc_status == 'proforma' else 'final'
    
    exact_date_time = now_j.strftime("%Y/%m/%d - %H:%M:%S")
    customer_name = request.form.get('customer_name', 'مشتری محترم').strip()
    customer_phone = request.form.get('customer_phone', '').strip()
    
    # ثبت یا دریافت مشتری در CRM
    customer = get_or_create_customer(customer_name, customer_phone)
    
    second_seller = request.form.get('second_seller_id')
    second_seller_id = int(second_seller) if second_seller else None
    
    total_amount = int(request.form.get('total_amount', '0').replace(',', ''))
    
    # مبالغ پرداخت ترکیبی
    paid_pos = int(request.form.get('paid_pos', '0').replace(',', '') or '0')
    paid_card = int(request.form.get('paid_card', '0').replace(',', '') or '0')
    paid_cash = int(request.form.get('paid_cash', '0').replace(',', '') or '0')
    
    # چک‌های صیادی چندگانه
    cheque_sayads = request.form.getlist('cheque_sayad[]')
    cheque_banks = request.form.getlist('cheque_bank[]')
    cheque_amounts = request.form.getlist('cheque_amount[]')
    cheque_due_dates = request.form.getlist('cheque_due_date[]')
    
    paid_cheque = 0
    cheques_to_create = []
    for idx in range(len(cheque_sayads)):
        sayad_val = cheque_sayads[idx].strip() if idx < len(cheque_sayads) else ''
        if sayad_val:
            raw_amt = cheque_amounts[idx].replace(',', '') if idx < len(cheque_amounts) and cheque_amounts[idx] else '0'
            chk_amt = int(raw_amt) if raw_amt else 0
            paid_cheque += chk_amt
            cheques_to_create.append({
                'sayad': sayad_val,
                'bank': cheque_banks[idx] if idx < len(cheque_banks) and cheque_banks[idx] else 'نامشخص',
                'amount': chk_amt,
                'due_date': cheque_due_dates[idx] if idx < len(cheque_due_dates) and cheque_due_dates[idx] else 'نامشخص'
            })

    # مانده تسویه نشده / بیعانه
    remaining_balance = int(request.form.get('remaining_balance', '0').replace(',', '') or '0')
    total_paid_immediate = paid_pos + paid_card + paid_cash
    
    if remaining_balance == 0 and (total_paid_immediate + paid_cheque) < total_amount:
        remaining_balance = total_amount - (total_paid_immediate + paid_cheque)

    is_settled = (remaining_balance <= 0) and (paid_cheque == 0)
    
    # تشخیص روش پرداخت برای نمایش در فاکتور
    active_methods = []
    if paid_pos > 0: active_methods.append(f"کارتخوان: {paid_pos:,}")
    if paid_card > 0: active_methods.append(f"کارت/شبا: {paid_card:,}")
    if paid_cash > 0: active_methods.append(f"نقد: {paid_cash:,}")
    if paid_cheque > 0: active_methods.append(f"چک صیادی: {paid_cheque:,}")
    if remaining_balance > 0: active_methods.append(f"مانده بیعانه: {remaining_balance:,}")
    payment_method_str = " | ".join(active_methods) if active_methods else "کارتخوان (POS)"
    
    invoice_number = request.form.get('invoice_number', '').strip()
    if not invoice_number:
        invoice_number = f"INV-{now_j.year}{now_j.month:02d}-{int(datetime.utcnow().timestamp()) % 10000}"

    dest_card = request.form.get('dest_card_number', '').strip()
    dest_sheba = request.form.get('dest_sheba_number', '').strip()

    new_inv = Invoice(
        invoice_number=invoice_number,
        customer_id=customer.id if customer else None,
        customer_name=customer_name,
        customer_phone=customer_phone,
        items_desc=request.form.get('items_desc'),
        status=status,
        proforma_valid_until=request.form.get('proforma_valid_until'),
        invoice_type=inv_type,
        return_reason=request.form.get('return_reason'),
        payment_method=payment_method_str,
        paid_pos=paid_pos,
        paid_card=paid_card,
        paid_cash=paid_cash,
        paid_cheque=paid_cheque,
        remaining_balance=remaining_balance,
        paid_amount=total_paid_immediate,
        dest_card_number=dest_card,
        dest_sheba_number=dest_sheba,
        payment_tracking_code=request.form.get('payment_tracking_code'),
        total_amount=total_amount,
        due_settlement_date=request.form.get('due_settlement_date'),
        is_settled=(remaining_balance <= 0),
        seller_id=session['user_id'],
        second_seller_id=second_seller_id,
        split_ratio=int(request.form.get('split_ratio', 100)),
        customer_rating=int(request.form.get('customer_rating', 5)),
        shamsi_year=now_j.year,
        shamsi_month=now_j.month,
        shamsi_day=now_j.day,
        shamsi_date_time=exact_date_time,
        shop_id=session.get('shop_id', 1)
    )
    db.session.add(new_inv)
    db.session.flush()

    # پردازش اقلام فاکتور و کسر از انبار
    inv_item_ids = request.form.getlist('item_inventory_id[]')
    custom_names = request.form.getlist('item_custom_name[]')
    custom_cats = request.form.getlist('item_category[]')
    quantities = request.form.getlist('item_quantity[]')
    prices = request.form.getlist('item_price[]')
    
    total_actual_buy_cost = 0
    categories_used = set()

    for idx in range(len(quantities)):
        qty = int(quantities[idx]) if idx < len(quantities) and quantities[idx] else 1
        raw_p = prices[idx].replace(',', '') if idx < len(prices) and prices[idx] else '0'
        price = int(raw_p) if raw_p else 0
        
        item_id_val = inv_item_ids[idx] if idx < len(inv_item_ids) and inv_item_ids[idx] else None
        inv_item = InventoryItem.query.get(int(item_id_val)) if item_id_val else None
        
        name_val = inv_item.name if inv_item else (custom_names[idx] if idx < len(custom_names) and custom_names[idx] else 'تجهیزات بهداشتی')
        cat_val = inv_item.category if inv_item else (custom_cats[idx] if idx < len(custom_cats) and custom_cats[idx] else 'عمومی')
        
        # ثبت خودکار دسته جدید در صورت نبود
        if cat_val and not Category.query.filter_by(name=cat_val).first():
            db.session.add(Category(name=cat_val))
            db.session.commit()
            
        categories_used.add(cat_val)
        buy_p = inv_item.buy_price if inv_item else int(price * 0.75)
        
        row_total = price * qty
        row_profit = row_total - (buy_p * qty)
        
        total_actual_buy_cost += (buy_p * qty)
        
        inv_row = InvoiceItem(
            invoice_id=new_inv.id,
            inventory_item_id=inv_item.id if inv_item else None,
            item_name=name_val,
            category=cat_val,
            quantity=qty,
            unit_buy_price=buy_p,
            unit_sell_price=price,
            total_price=row_total,
            row_profit=row_profit
        )
        db.session.add(inv_row)
        
        # کسر از انبار برای فاکتور قطعی
        if status == 'final' and inv_item:
            if inv_type == 'sale':
                record_stock_change(inv_item.id, session.get('shop_id', 1), 'sale', -qty, new_inv.invoice_number, session.get('full_name'), f"فروش در فاکتور {new_inv.invoice_number}")
            elif inv_type == 'return':
                record_stock_change(inv_item.id, session.get('shop_id', 1), 'return', qty, new_inv.invoice_number, session.get('full_name'), f"مرجوعی فاکتور {new_inv.invoice_number}")

    new_inv.categories_json = json.dumps(list(categories_used), ensure_ascii=False)
    new_inv.actual_buy_cost = total_actual_buy_cost if total_actual_buy_cost > 0 else int(total_amount * 0.75)
    new_inv.real_profit = total_amount - new_inv.actual_buy_cost

    # ثبت چک‌های صیادی ایجاد شده
    for chk_data in cheques_to_create:
        chk = Cheque(
            invoice_id=new_inv.id,
            sayad_number=chk_data['sayad'],
            bank_name=chk_data['bank'],
            customer_name=new_inv.customer_name,
            customer_phone=new_inv.customer_phone,
            amount=chk_data['amount'],
            due_shamsi_date=chk_data['due_date'],
            shop_id=session.get('shop_id', 1),
            status='pending'
        )
        db.session.add(chk)

    # ثبت در CRM مشتری
    if customer:
        customer.total_purchases += total_amount
        if remaining_balance > 0:
            customer.outstanding_balance += remaining_balance

    db.session.commit()
    
    log_activity(f"ثبت سند {new_inv.invoice_number} ({status}) به مبلغ {total_amount:,} تومان با پرداخت ترکیبی", session.get('full_name'), "فروش")
    flash('فاکتور با پرداخت ترکیبی و کسر دقیق از انبار ثبت گردید.', 'success')
    return redirect(url_for('seller_dashboard'))

@app.route('/invoice/convert/<int:invoice_id>', methods=['POST'])
def convert_proforma(invoice_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    inv = Invoice.query.get_or_404(invoice_id)
    inv.status = 'final'
    
    # کسر اقلام از انبار در لحظه قطعی شدن
    for row in inv.items:
        if row.inventory_item_id:
            record_stock_change(row.inventory_item_id, inv.shop_id, 'sale', -row.quantity, inv.invoice_number, session.get('full_name'), f"تبدیل پیش‌فاکتور به قطعی {inv.invoice_number}")

    db.session.commit()
    log_activity(f"تبدیل پیش‌فاکتور {inv.invoice_number} به فاکتور قطعی و کسر انبار", session.get('full_name'), "فروش")
    flash(f'پیش‌فاکتور {inv.invoice_number} به فاکتور قطعی تبدیل و از انبار کسر شد.', 'success')
    return redirect(url_for('seller_dashboard'))

@app.route('/invoice/settle_deposit/<int:invoice_id>', methods=['POST'])
def settle_deposit(invoice_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    inv = Invoice.query.get_or_404(invoice_id)
    inv.paid_amount = inv.total_amount
    inv.is_settled = True
    inv.payment_method = 'pos'
    if inv.customer and inv.customer.outstanding_balance > 0:
        inv.customer.outstanding_balance = max(inv.customer.outstanding_balance - (inv.total_amount - inv.paid_amount), 0)
    db.session.commit()
    log_activity(f"تسویه کامل مانده فاکتور {inv.invoice_number}", session.get('full_name'), "مالی")
    flash(f'مانده فاکتور {inv.invoice_number} به طور کامل تسویه شد.', 'success')
    return redirect(url_for('seller_dashboard'))

# ==================== چاپ فاکتورها (A4 و فیش پرینتر) ====================
@app.route('/invoice/print/a4/<int:invoice_id>')
def print_invoice_a4(invoice_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    invoice = Invoice.query.get_or_404(invoice_id)
    return render_template('print_a4.html', invoice=invoice)

@app.route('/invoice/print/thermal/<int:invoice_id>')
def print_invoice_thermal(invoice_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    invoice = Invoice.query.get_or_404(invoice_id)
    return render_template('print_thermal.html', invoice=invoice)

# ==================== داشبورد مدیریت کل ====================
@app.route('/admin')
def admin_dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    now_j = jdatetime.datetime.now()
    selected_month = request.args.get('month', default=now_j.month, type=int)
    search_query = request.args.get('search', '').strip()
    seller_filter = request.args.get('seller_filter', '').strip()
    
    settings = Settings.query.first()
    sellers = User.query.filter_by(role='seller', is_active=True).all()
    
    sellers_data = []
    shop1_total = 0
    shop2_total = 0
    total_commissions = 0
    total_sales_all = 0
    estimated_gross_profit = 0
    
    chart_sellers_labels = []
    chart_sellers_data = []
    
    for s in sellers:
        s_stats = calculate_seller_exact_stats(s.id, now_j.year, selected_month, s.commission_rate, settings)
        total_commissions += s_stats['settled_commission']
        total_sales_all += s_stats['net_sales']
        estimated_gross_profit += s_stats['real_profit_share']
        
        if s.shop_id == 1:
            shop1_total += s_stats['net_sales']
        else:
            shop2_total += s_stats['net_sales']
            
        sellers_data.append({'user': s, 'stats': s_stats})
        chart_sellers_labels.append(s.full_name)
        chart_sellers_data.append(s_stats['net_sales'])
        
    expenses = Expense.query.filter_by(shamsi_year=now_j.year, shamsi_month=selected_month).all()
    total_expenses = sum(e.amount for e in expenses)
    
    petty_deposits = PettyCashDeposit.query.filter_by(shamsi_year=now_j.year, shamsi_month=selected_month).all()
    total_petty_deposits = sum(d.amount for d in petty_deposits)
    
    all_categories = Category.query.all()
    all_month_items = InvoiceItem.query.join(Invoice).filter(Invoice.shamsi_year == now_j.year, Invoice.shamsi_month == selected_month, Invoice.status == 'final').all()
    cat_counts = {cat.name: 0 for cat in all_categories}
    for item in all_month_items:
        if item.category in cat_counts:
            cat_counts[item.category] += item.quantity
            
    chart_category_labels = list(cat_counts.keys())
    chart_category_data = list(cat_counts.values())

    cheques = Cheque.query.order_by(Cheque.id.desc()).all()
    pending_cheques = [c for c in cheques if c.status == 'pending']
    pending_cheques_total = sum(c.amount for c in pending_cheques)
    
    all_inventory = InventoryItem.query.all()
    low_stock_count = len([i for i in all_inventory if i.stock_quantity <= i.min_alert_stock])
    
    query = Invoice.query.filter_by(shamsi_year=now_j.year, shamsi_month=selected_month)
    if seller_filter:
        query = query.filter_by(seller_id=int(seller_filter))
        
    if search_query:
        query = query.filter(
            (Invoice.customer_name.contains(search_query)) |
            (Invoice.customer_phone.contains(search_query)) |
            (Invoice.invoice_number.contains(search_query))
        )
    all_invoices_raw = query.order_by(Invoice.created_at.desc()).all()
    
    all_invoices = []
    for inv in all_invoices_raw:
        all_invoices.append({
            'id': inv.id,
            'invoice_number': inv.invoice_number,
            'status': inv.status,
            'invoice_type': inv.invoice_type,
            'return_reason': inv.return_reason,
            'shop_name': inv.shop.name if inv.shop else '',
            'seller_name': inv.seller.full_name if inv.seller else '',
            'second_seller_id': inv.second_seller_id,
            'split_ratio': inv.split_ratio,
            'customer_name': inv.customer_name,
            'customer_phone': inv.customer_phone,
            'total_amount': inv.total_amount,
            'paid_amount': inv.paid_amount,
            'payment_method': inv.payment_method,
            'dest_card_number': inv.dest_card_number,
            'payment_tracking_code': inv.payment_tracking_code,
            'cheque_sayad': inv.cheques[0].sayad_number if inv.cheques else '',
            'cheque_bank': inv.cheques[0].bank_name if inv.cheques else '',
            'cheque_due_date': inv.cheques[0].due_shamsi_date if inv.cheques else '',
            'due_settlement_date': inv.due_settlement_date,
            'shamsi_date_time': inv.shamsi_date_time,
            'items_desc': inv.items_desc or '',
            'categories_json': inv.categories_json or ''
        })
    
    shops = Shop.query.all()
    logs = AuditLog.query.order_by(AuditLog.id.desc()).limit(25).all()
    bank_accounts = BankAccount.query.order_by(BankAccount.id.desc()).all()
    
    return render_template(
        'admin_dashboard.html',
        sellers_data=sellers_data,
        all_sellers=sellers,
        seller_filter=seller_filter,
        shop1_total=shop1_total,
        shop2_total=shop2_total,
        total_commissions=total_commissions,
        total_expenses=total_expenses,
        total_petty_deposits=total_petty_deposits,
        petty_deposits=petty_deposits,
        expenses=expenses,
        estimated_gross_profit=estimated_gross_profit,
        cheques=cheques,
        pending_cheques_count=len(pending_cheques),
        pending_cheques_total=pending_cheques_total,
        all_inventory=all_inventory,
        low_stock_count=low_stock_count,
        months=PERSIAN_MONTHS,
        selected_month=selected_month,
        selected_month_name=PERSIAN_MONTHS.get(selected_month, ''),
        current_year=now_j.year,
        shops=shops,
        all_categories=all_categories,
        bank_accounts=bank_accounts,
        chart_sellers_labels=chart_sellers_labels,
        chart_sellers_data=chart_sellers_data,
        chart_category_labels=chart_category_labels,
        chart_category_data=chart_category_data,
        all_invoices=all_invoices,
        search_query=search_query,
        settings=settings,
        logs=logs
    )

# ==================== مدیریت حساب‌های بانکی (شماره کارت و شماره شبا) ====================
@app.route('/admin/bank_account/add', methods=['POST'])
def add_bank_account():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    title = request.form.get('title', 'حساب بانکی طهماسبی').strip()
    bank_name = request.form.get('bank_name', 'ملی').strip()
    account_owner = request.form.get('account_owner', 'طهماسبی').strip()
    account_type = request.form.get('account_type', 'both')
    card_number = request.form.get('card_number', '').strip()
    sheba_number = request.form.get('sheba_number', '').strip()

    if sheba_number and not sheba_number.upper().startswith('IR'):
        sheba_number = 'IR' + sheba_number

    acc = BankAccount(
        title=title,
        bank_name=bank_name,
        account_owner=account_owner,
        account_type=account_type,
        card_number=card_number,
        sheba_number=sheba_number,
        is_active=True
    )
    db.session.add(acc)
    db.session.commit()
    
    log_activity(f"افزودن حساب/کارت بانکی {title} ({bank_name})", session.get('full_name'), "تنظیمات")
    flash('حساب بانکی جدید با موفقیت به سیستم اضافه شد.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/bank_account/delete/<int:account_id>', methods=['POST'])
def delete_bank_account(account_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    acc = BankAccount.query.get_or_404(account_id)
    title = acc.title
    db.session.delete(acc)
    db.session.commit()
    
    log_activity(f"حذف حساب بانکی {title}", session.get('full_name'), "تنظیمات")
    flash(f'حساب بانکی {title} حذف گردید.', 'warning')
    return redirect(url_for('admin_dashboard'))

# ==================== ماژول انبارداری و انبارگردانی ====================
@app.route('/inventory')
def inventory_view():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    all_inventory = InventoryItem.query.all()
    shops = Shop.query.all()
    all_categories = Category.query.all()
    transfers = StockTransfer.query.order_by(StockTransfer.id.desc()).limit(15).all()
    ai_insights = get_inventory_ai_insights()

    return render_template(
        'inventory.html',
        all_inventory=all_inventory,
        shops=shops,
        all_categories=all_categories,
        transfers=transfers,
        ai_insights=ai_insights
    )

@app.route('/admin/inventory/add', methods=['POST'])
def add_inventory_item():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    category_val = request.form.get('category')
    if category_val == '__custom__':
        category_val = request.form.get('custom_category_name', '').strip()
        if not category_val:
            category_val = 'سایر و متفرقه'
        if not Category.query.filter_by(name=category_val).first():
            db.session.add(Category(name=category_val))
            db.session.commit()
            
    buy_raw = request.form.get('buy_price', '').replace(',', '')
    sell_raw = request.form.get('sell_price', '').replace(',', '')
    item = InventoryItem(
        name=request.form.get('name'),
        category=category_val,
        shop_id=int(request.form.get('shop_id', 1)),
        stock_quantity=int(request.form.get('stock_quantity', 5)),
        min_alert_stock=int(request.form.get('min_alert_stock', 2)),
        buy_price=int(buy_raw) if buy_raw else 0,
        sell_price=int(sell_raw) if sell_raw else 0
    )
    db.session.add(item)
    db.session.commit()
    
    record_stock_change(item.id, item.shop_id, 'purchase_in', item.stock_quantity, 'INIT', session.get('full_name'), "موجودی اولیه")
    log_activity(f"افزودن کالای {item.name} به انبار", session.get('full_name'), "انبار")
    flash(f'کالای {item.name} با موفقیت در انبار ثبت گردید.', 'success')
    return redirect(url_for('inventory_view'))

@app.route('/admin/inventory/edit/<int:item_id>', methods=['POST'])
def edit_inventory_item(item_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    item = InventoryItem.query.get_or_404(item_id)
    item.name = request.form.get('name')
    item.category = request.form.get('category')
    item.shop_id = int(request.form.get('shop_id'))
    old_qty = item.stock_quantity
    new_qty = int(request.form.get('stock_quantity', 0))
    item.stock_quantity = new_qty
    item.min_alert_stock = int(request.form.get('min_alert_stock', 2))
    
    buy_raw = request.form.get('buy_price', '').replace(',', '')
    sell_raw = request.form.get('sell_price', '').replace(',', '')
    item.buy_price = int(buy_raw) if buy_raw else 0
    item.sell_price = int(sell_raw) if sell_raw else 0
    
    if new_qty != old_qty:
        record_stock_change(item.id, item.shop_id, 'adjustment', new_qty - old_qty, 'MANUAL_EDIT', session.get('full_name'), "ویرایش دستی انبار")

    db.session.commit()
    log_activity(f"ویرایش کالای {item.name} در انبار", session.get('full_name'), "انبار")
    flash(f'اطلاعات کالای {item.name} به‌روزرسانی شد.', 'success')
    return redirect(url_for('inventory_view'))

@app.route('/admin/inventory/delete/<int:item_id>', methods=['POST'])
def delete_inventory_item(item_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    item = InventoryItem.query.get_or_404(item_id)
    item_name = item.name
    db.session.delete(item)
    db.session.commit()
    log_activity(f"حذف کالای {item_name} از انبار", session.get('full_name'), "انبار")
    flash(f'کالای {item_name} از انبار حذف گردید.', 'success')
    return redirect(url_for('inventory_view'))

@app.route('/admin/inventory/stocktaking', methods=['POST'])
def stocktaking_adjust():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    item_id = int(request.form.get('item_id'))
    actual_stock = int(request.form.get('actual_stock', 0))
    reason = request.form.get('reason', 'انبارگردانی دوره‌ای')
    
    item = InventoryItem.query.get_or_404(item_id)
    diff = actual_stock - item.stock_quantity
    item.stock_quantity = actual_stock
    
    record_stock_change(item.id, item.shop_id, 'adjustment', diff, 'STOCKTAKING', session.get('full_name'), reason)
    log_activity(f"انبارگردانی کالای {item.name}: موجودی جدید {actual_stock} ({diff:+d})", session.get('full_name'), "انبار")
    flash(f'انبارگردانی کالای {item.name} با موفقیت ثبت شد.', 'success')
    return redirect(url_for('inventory_view'))

@app.route('/transfer/request', methods=['POST'])
def request_transfer():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    now_j = jdatetime.datetime.now().strftime("%Y/%m/%d - %H:%M:%S")
    st = StockTransfer(
        item_name=request.form.get('item_name'),
        from_shop_id=int(request.form.get('from_shop_id')),
        to_shop_id=session.get('shop_id', 1),
        quantity=int(request.form.get('quantity', 1)),
        requested_by=session['full_name'],
        shamsi_date_time=now_j
    )
    db.session.add(st)
    db.session.commit()
    log_activity(f"درخواست انتقال {st.quantity} عدد {st.item_name} بین شعب", session.get('full_name'), "انبار")
    flash('درخواست انتقال کالا با موفقیت ثبت شد.', 'success')
    return redirect(url_for('inventory_view'))

# ==================== ماژول حقوق، دستمزد و مساعده ====================
@app.route('/admin/payroll')
def payroll_view():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    
    now_j = jdatetime.datetime.now()
    selected_month = request.args.get('month', default=now_j.month, type=int)
    settings = Settings.query.first()
    sellers = User.query.filter_by(is_active=True).all()
    
    payroll_items = []
    total_payroll_payable = 0
    total_commissions_settled = 0
    total_tier_bonuses = 0

    for u in sellers:
        stats = calculate_seller_exact_stats(u.id, now_j.year, selected_month, u.commission_rate, settings)
        
        # محاسبه مساعده‌های ثبت شده ماه
        adv_expenses = Expense.query.filter_by(
            shamsi_year=now_j.year,
            shamsi_month=selected_month,
            category='مساعده'
        ).filter(Expense.title.contains(u.full_name)).all()
        advances = sum(e.amount for e in adv_expenses)
        
        base_sal = u.base_salary or 0
        final_payable = base_sal + stats['settled_commission'] + stats['tier_bonus_amount'] - advances
        
        total_payroll_payable += final_payable
        total_commissions_settled += stats['settled_commission']
        total_tier_bonuses += stats['tier_bonus_amount']

        payroll_items.append({
            'user': u,
            'stats': stats,
            'base_salary': base_sal,
            'advances': advances,
            'final_payable': final_payable
        })

    return render_template(
        'payroll.html',
        payroll_items=payroll_items,
        months=PERSIAN_MONTHS,
        selected_month=selected_month,
        current_month_name=PERSIAN_MONTHS.get(selected_month, ''),
        current_year=now_j.year,
        total_payroll_payable=total_payroll_payable,
        total_commissions_settled=total_commissions_settled,
        total_tier_bonuses=total_tier_bonuses
    )

@app.route('/admin/payroll/base_salary/<int:user_id>', methods=['POST'])
def update_user_base_salary(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    user = User.query.get_or_404(user_id)
    month = request.form.get('month', 1)
    base_sal_raw = request.form.get('base_salary', '0').replace(',', '')
    user.base_salary = int(base_sal_raw) if base_sal_raw else 0
    db.session.commit()
    log_activity(f"تغییر حقوق پایه {user.full_name} به {user.base_salary:,} تومان", session.get('full_name'), "حقوق")
    flash(f'حقوق پایه {user.full_name} به‌روزرسانی شد.', 'success')
    return redirect(url_for('payroll_view', month=month))

@app.route('/admin/payroll/advance/<int:user_id>', methods=['POST'])
def add_user_advance(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    user = User.query.get_or_404(user_id)
    month = int(request.form.get('month', 1))
    now_j = jdatetime.datetime.now()
    adv_raw = request.form.get('advance_amount', '0').replace(',', '')
    amount = int(adv_raw) if adv_raw else 0
    notes = request.form.get('notes', '')
    
    exp = Expense(
        title=f"مساعده {user.full_name} ({notes})",
        amount=amount,
        category='مساعده',
        shamsi_year=now_j.year,
        shamsi_month=month,
        shamsi_date_time=now_j.strftime("%Y/%m/%d - %H:%M:%S"),
        shop_id=user.shop_id or 1,
        created_by=session.get('full_name')
    )
    db.session.add(exp)
    db.session.commit()
    log_activity(f"ثبت مساعده {amount:,} تومان برای {user.full_name}", session.get('full_name'), "حقوق")
    flash(f'مساعده {user.full_name} با موفقیت ثبت گردید.', 'success')
    return redirect(url_for('payroll_view', month=month))

# ==================== ماژول مشتریان CRM ====================
@app.route('/admin/customers')
def customers_view():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    customers = Customer.query.order_by(Customer.total_purchases.desc()).all()
    return render_template('customers.html', customers=customers)

# ==================== API هوش مصنوعی صدور فاکتور ====================
@app.route('/api/ai/parse_invoice', methods=['POST'])
def api_parse_invoice():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'احراز هویت نشده'}), 401
    
    data = request.get_json() or {}
    raw_text = data.get('text', '')
    res = parse_smart_invoice_text(raw_text, session.get('shop_id', 1))
    return jsonify(res)

# ==================== سایر عملیات مدیریت (تنخواه، کمیسیون، تنظیمات) ====================
@app.route('/admin/commission/<int:user_id>', methods=['POST'])
def update_commission(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    user = User.query.get_or_404(user_id)
    month = request.form.get('month', 1)
    rate = float(request.form.get('commission_rate', 1.0))
    if 0.01 <= rate <= 5.0:
        user.commission_rate = round(rate, 2)
        db.session.commit()
        log_activity(f"تغییر درصد پایه {user.full_name} به {user.commission_rate}%", session.get('full_name'), "حقوق")
        flash(f'درصد پایه {user.full_name} به {user.commission_rate}% تغییر یافت.', 'success')
    return redirect(url_for('admin_dashboard', month=month))

@app.route('/admin/user/set_password/<int:user_id>', methods=['POST'])
def set_custom_password(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    user = User.query.get_or_404(user_id)
    user.set_password(request.form.get('new_password'))
    db.session.commit()
    log_activity(f"تغییر رمز کاربر {user.full_name}", session.get('full_name'), "امنیت")
    flash(f'رمز عبور جدید برای {user.full_name} ثبت شد.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/change_my_password', methods=['POST'])
def admin_change_own_password():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    user = User.query.get_or_404(session['user_id'])
    new_pw = request.form.get('new_admin_password')
    if new_pw:
        user.set_password(new_pw)
        db.session.commit()
        log_activity("تغییر رمز عبور ورود توسط مدیر کل", user.full_name, "امنیت")
        flash('رمز عبور مدیریت با موفقیت تغییر یافت.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/user/delete/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    user = User.query.get_or_404(user_id)
    user.is_active = False
    db.session.commit()
    log_activity(f"حذف پرسنل {user.full_name}", session.get('full_name'), "پرسنل")
    flash(f'پرسنل {user.full_name} غیرفعال گردید.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/invoice/delete/<int:invoice_id>', methods=['POST'])
def delete_invoice(invoice_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    inv = Invoice.query.get_or_404(invoice_id)
    inv_num = inv.invoice_number
    db.session.delete(inv)
    db.session.commit()
    log_activity(f"حذف فاکتور شماره {inv_num}", session.get('full_name'), "فروش")
    flash(f'فاکتور شماره {inv_num} حذف شد.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/settings/update', methods=['POST'])
def update_settings():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    settings = Settings.query.first()
    settings.tier1_min = int(request.form.get('tier1_min', '0').replace(',', ''))
    settings.tier1_bonus = float(request.form.get('tier1_bonus', 0.25))
    settings.tier2_min = int(request.form.get('tier2_min', '0').replace(',', ''))
    settings.tier2_bonus = float(request.form.get('tier2_bonus', 0.50))
    db.session.commit()
    log_activity("به‌روزرسانی پله‌های تارگت پورسانت", session.get('full_name'), "مالی")
    flash('تنظیمات پله‌های تارگت پورسانت به‌روزرسانی شد.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/user/add', methods=['POST'])
def add_user():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    username = request.form.get('username', '').strip()
    if User.query.filter_by(username=username).first():
        flash('این نام کاربری تکراری است.', 'error')
        return redirect(url_for('admin_dashboard'))
    new_user = User(
        username=username,
        full_name=request.form.get('full_name'),
        role='seller',
        shop_id=int(request.form.get('shop_id', 1)),
        commission_rate=float(request.form.get('commission_rate', 1.0))
    )
    new_user.set_password(request.form.get('password'))
    db.session.add(new_user)
    db.session.commit()
    log_activity(f"تعریف پرسنل جدید ({new_user.full_name})", session.get('full_name'), "پرسنل")
    flash('پرسنل جدید با موفقیت ثبت شد.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/cheque/status/<int:cheque_id>', methods=['POST'])
def update_cheque_status(cheque_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    chk = Cheque.query.get_or_404(cheque_id)
    chk.status = request.form.get('status')
    if chk.status == 'passed':
        chk.passed_shamsi_date = jdatetime.datetime.now().strftime("%Y/%m/%d")
    db.session.commit()
    flash('وضعیت چک با موفقیت به‌روزرسانی شد.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/petty_deposit/add', methods=['POST'])
def add_petty_deposit():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    now_j = jdatetime.datetime.now()
    amount = int(request.form.get('amount', '0').replace(',', ''))
    dep = PettyCashDeposit(
        title=request.form.get('title'),
        amount=amount,
        shop_id=int(request.form.get('shop_id', 1)),
        shamsi_year=now_j.year,
        shamsi_month=now_j.month,
        shamsi_date_time=now_j.strftime("%Y/%m/%d - %H:%M:%S"),
        created_by=session['full_name']
    )
    db.session.add(dep)
    db.session.commit()
    log_activity(f"شارژ تنخواه به مبلغ {amount:,} تومان", session.get('full_name'), "مالی")
    flash('شارژ تنخواه با موفقیت ثبت شد.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/expense/add', methods=['POST'])
def add_expense():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    now_j = jdatetime.datetime.now()
    amount = int(request.form.get('amount', '0').replace(',', ''))
    exp = Expense(
        title=request.form.get('title'),
        amount=amount,
        category=request.form.get('category'),
        shamsi_year=now_j.year,
        shamsi_month=now_j.month,
        shamsi_date_time=now_j.strftime("%Y/%m/%d - %H:%M:%S"),
        shop_id=session.get('shop_id', 1),
        created_by=session['full_name']
    )
    db.session.add(exp)
    db.session.commit()
    log_activity(f"ثبت هزینه تنخواه {exp.title} به مبلغ {exp.amount:,} تومان", session.get('full_name'), "مالی")
    flash('هزینه در تنخواه ثبت شد.', 'success')
    return redirect(url_for('index'))

@app.route('/admin/backup')
def download_backup():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    now_str = jdatetime.datetime.now().strftime("%Y%m%d_%H%M")
    log_activity("دانلود فایل بکاپ دیتابیس", session.get('full_name'), "امنیت")
    return send_file(db_path, as_attachment=True, download_name=f"Backup_Tahmasebi_{now_str}.db")

@app.route('/admin/export/excel')
def export_excel():
    if 'user_id' not in session or session.get('role') != 'admin':
        return redirect(url_for('login'))
    now_j = jdatetime.datetime.now()
    month = request.args.get('month', default=now_j.month, type=int)
    month_name = PERSIAN_MONTHS.get(month, '')
    settings = Settings.query.first()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"پورسانت {month_name}"
    ws.views.sheetView[0].rightToLeft = True

    headers = ['نام پرسنل', 'شعبه طهماسبی', 'تعداد اسناد', 'فروش خالص (تومان)', 'درصد پایه', 'درصد با پله تارگت', 'مبلغ پورسانت قطعی (تومان)']
    ws.append(headers)

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sellers = User.query.filter_by(role='seller', is_active=True).all()
    for s in sellers:
        s_stats = calculate_seller_exact_stats(s.id, now_j.year, month, s.commission_rate, settings)
        ws.append([s.full_name, s.shop.name if s.shop else '', s_stats['sales_count'], s_stats['net_sales'], f"{s.commission_rate}%", f"{s_stats['effective_rate']}%", s_stats['settled_commission']])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Tahmasebi_Report_{month_name}_{now_j.year}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ==================== مقداردهی اولیه و ارتقای خودکار اسکیمای دیتابیس ====================
def auto_migrate_db():
    import sqlite3
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # لیست ستون‌های جدید مورد نیاز در جداول
    migrations = [
        ("settings", "store_name", "TEXT DEFAULT 'مجموعه فروشگاه‌های تخصصی طهماسبی'"),
        ("settings", "store_phone", "TEXT DEFAULT '021-12345678'"),
        ("settings", "store_warranty_text", "TEXT DEFAULT 'کلیه اقلام دارای گارانتی اصالت کالا می‌باشند.'"),
        ("users", "base_salary", "BIGINT DEFAULT 0"),
        ("users", "phone", "TEXT"),
        ("users", "card_number", "TEXT"),
        ("users", "created_at", "DATETIME"),
        ("shops", "phone", "TEXT"),
        ("shops", "address", "TEXT"),
        ("categories", "icon", "TEXT DEFAULT '📦'"),
        ("invoices", "customer_id", "INTEGER"),
        ("invoices", "subtotal_amount", "BIGINT DEFAULT 0"),
        ("invoices", "discount_amount", "BIGINT DEFAULT 0"),
        ("invoices", "actual_buy_cost", "BIGINT DEFAULT 0"),
        ("invoices", "real_profit", "BIGINT DEFAULT 0"),
        ("invoices", "is_settled", "BOOLEAN DEFAULT 1"),
        ("invoices", "shamsi_day", "INTEGER"),
        ("invoices", "paid_pos", "BIGINT DEFAULT 0"),
        ("invoices", "paid_card", "BIGINT DEFAULT 0"),
        ("invoices", "paid_cash", "BIGINT DEFAULT 0"),
        ("invoices", "paid_cheque", "BIGINT DEFAULT 0"),
        ("invoices", "remaining_balance", "BIGINT DEFAULT 0"),
        ("invoices", "dest_sheba_number", "TEXT")
    ]
    
    for table, col, col_def in migrations:
        try:
            cursor.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_def}")
            conn.commit()
        except Exception:
            pass # ستون از قبل وجود دارد
            
    conn.close()

with app.app_context():
    db.create_all()
    auto_migrate_db()
    
    if not BankAccount.query.first():
        acc1 = BankAccount(
            title='کارت اصلی فروشگاه طهماسبی',
            bank_name='بانک ملی ایران',
            account_owner='ابوالفضل طهماسبی',
            account_type='both',
            card_number='6037997512345678',
            sheba_number='IR120170000000123456789012'
        )
        acc2 = BankAccount(
            title='حساب پایا / شبا شعبه ۲',
            bank_name='بانک ملت',
            account_owner='ابوالفضل طهماسبی',
            account_type='sheba',
            card_number='',
            sheba_number='IR980120000000987654321098'
        )
        db.session.add_all([acc1, acc2])
        db.session.commit()
    
    for cat_name in DEFAULT_CATEGORIES:
        if not Category.query.filter_by(name=cat_name).first():
            db.session.add(Category(name=cat_name))
    db.session.commit()
    
    if not Settings.query.first():
        db.session.add(Settings())
        db.session.commit()
        
    if not Shop.query.first():
        shop1 = Shop(name='فروشگاه طهماسبی - شعبه ۱ (مرکزی)', phone='021-11111111')
        shop2 = Shop(name='فروشگاه طهماسبی - شعبه ۲', phone='021-22222222')
        db.session.add_all([shop1, shop2])
        db.session.commit()

        admin = User(username='admin', full_name='مدیریت کل (طهماسبی)', role='admin', base_salary=0)
        admin.set_password('admin123')
        db.session.add(admin)

        u1 = User(username='naqdi', full_name='خانم نقدی', role='seller', shop_id=shop1.id, commission_rate=1.0, base_salary=15000000)
        u1.set_password('123456')
        u2 = User(username='amiri', full_name='خانم امیری', role='seller', shop_id=shop1.id, commission_rate=1.0, base_salary=15000000)
        u2.set_password('123456')
        u3 = User(username='bidrigh', full_name='خانم بیدریغ', role='seller', shop_id=shop2.id, commission_rate=1.2, base_salary=15000000)
        u3.set_password('123456')
        u4 = User(username='hajilou', full_name='خانم حاجیلو', role='seller', shop_id=shop2.id, commission_rate=0.8, base_salary=15000000)
        u4.set_password('123456')
        u5 = User(username='zamani', full_name='ابوالفضل زمانی', role='seller', shop_id=shop1.id, commission_rate=5.0, base_salary=20000000)
        u5.set_password('123456')
        db.session.add_all([u1, u2, u3, u4, u5])
        db.session.commit()

        items = [
            InventoryItem(name='هود داتیس مدل 522 مخفی', category='هود', shop_id=shop1.id, stock_quantity=8, min_alert_stock=2, buy_price=6500000, sell_price=8900000),
            InventoryItem(name='گاز 5 شعله اخوان مدل GI-135', category='گاز صفحه‌ای', shop_id=shop1.id, stock_quantity=10, min_alert_stock=2, buy_price=7200000, sell_price=9800000),
            InventoryItem(name='سینک گرانیتی فونیکس دو لگن', category='سینک', shop_id=shop2.id, stock_quantity=6, min_alert_stock=2, buy_price=5400000, sell_price=7500000),
            InventoryItem(name='روشویی کابینتی ضدآب فول‌ست', category='روشویی کابینتی', shop_id=shop2.id, stock_quantity=4, min_alert_stock=2, buy_price=4200000, sell_price=6800000)
        ]
        db.session.add_all(items)
        db.session.commit()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)